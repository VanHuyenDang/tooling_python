import tkinter as tk
from tkinter import ttk
import math
import openpyxl
import os.path
import numpy as np
from common_libs import scan_excel
from common_libs import adjust_excel_format
from common_libs import check_excel_file, xuly_tinh_toan_excel
from common_libs import scan_excel_to_show

def refresh(param1_combobox, 
            param2_combobox,
            file_path
        ):
    ten_sp = param1_combobox.get()
    ma_sp = param2_combobox.get()
    # Process the selected parameters here
    thong_tin_ten_sp, thong_tin_ma_sp, thong_tin_gia_sp, thong_tin_sl_sp = scan_excel(file_path)
    param1_combobox['values'] = ()
    param2_combobox['values'] = ()
    param1_combobox['values'] = thong_tin_ten_sp
    param2_combobox['values'] = thong_tin_ma_sp

def tinh_tong_tien_nhap(file_path, tien_nhap_hang_display, doanh_thu_hang_display):
    tien_nhap_hang, doanh_so_ban = xuly_tinh_toan_excel(file_path)
    tien_nhap_hang_display.config(text=tien_nhap_hang)
    doanh_thu_hang_display.config(text=doanh_so_ban)

def hthi_thong_tin_sp(file_path, 
            param1_combobox, 
            param2_combobox,
            gia_ban_display,
            sl_nhap_display,
            sl_ban_ra_display,
            sl_con_lai_display
        ):
    ten_sp_tim = param1_combobox.get()
    ma_sp_tim = param2_combobox.get()
    # Process the selected parameters here
    sl_sp, gia_nhap_sp, gia_ban_ra, sl_sp_ban_ra, sl_sp_ton_kho = scan_excel_to_show(file_path, ma_sp_tim, ten_sp_tim)
    gia_ban_display.config(text=gia_ban_ra)
    sl_nhap_display.config(text=sl_sp)
    sl_ban_ra_display.config(text=sl_sp_ban_ra)
    sl_con_lai_display.config(text=sl_sp_ton_kho)
def cap_nhat(workbook, ten_hang_hoa, ma_san_pham, gia_tien_nhap, gia_ban, so_luong):
    # sheet = check_excel_file(file_path)
    sheet = workbook.active
    # Find the column index based on the header names
    code_column = None
    numbers_column = None
    for cell in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        for idx, header in enumerate(cell):
            if header == 'Mã sản phẩm':
                code_column = idx + 1
            elif header == 'Số lượng':
                numbers_column = idx + 1
            if code_column is not None and numbers_column is not None:
                break      
    # Find the row index based on code sản phẩm
    for idx, row in enumerate(sheet.iter_rows(), start=1):
        code_cell = row[code_column - 1]  # Assuming code_column is 1-indexed
        if code_cell.value == ma_san_pham:
            row_index = idx
            break
    # Scan the rows to find the product code
    product_found = False
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[code_column - 1] == ma_san_pham:
            product_found = True
            current_numbers = row[numbers_column - 1]
            new_numbers = int(current_numbers) + int(so_luong)
            sheet.cell(row=row_index, column=numbers_column, value=new_numbers)
            break

    # Add a new row if product code is not found
    if not product_found:
        sheet.append([str(ten_hang_hoa), str(ma_san_pham), float(gia_tien_nhap), int(so_luong), float(gia_ban)])
    # workbook.save(file_path)

def cap_nhat_ban(workbook, ten_hang_hoa_ban_ra, ma_san_pham_ban_ra, so_luong_ban_ra):
    sheet = workbook.active
    code_column = None
    numbers_column = None
    for cell in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        for idx, header in enumerate(cell):
            if header == 'Mã sản phẩm' or header == "Tên sản phẩm":
                code_column = idx + 1
            elif header == 'Số SP bán ra':
                numbers_column = idx + 1
            if code_column is not None and numbers_column is not None:
                break      
    # Find the row index based on code sản phẩm
    for idx, row in enumerate(sheet.iter_rows(), start=1):
        code_cell = row[code_column - 1]  
        if code_cell.value == ma_san_pham_ban_ra:
            row_index = idx
            break
    # Scan the rows to find the product code
    product_found = False
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[code_column - 1] == ma_san_pham_ban_ra:
            product_found = True
            current_numbers = row[numbers_column - 1]
            if current_numbers is None:
                current_numbers = int(0)
            new_numbers = current_numbers + int(so_luong_ban_ra)
            sheet.cell(row=row_index, column=numbers_column, value=new_numbers)
            break

    # Add a new row if product code is not found
    if not product_found:
        print("Bán sản phẩm không có trong cửa hàng")
    # workbook.save(file_path)


def nhap_hang(file_path, ten_hang_hoa_entry, ma_hang_hoa_entry, gia_tien_nhap_entry, so_luong_entry, gia_ban_1sp_entry):
    ten_hang_hoa = ten_hang_hoa_entry.get()
    ma_san_pham = ma_hang_hoa_entry.get()
    gia_tien_nhap = gia_tien_nhap_entry.get()
    so_luong = so_luong_entry.get()
    gia_ban = gia_ban_1sp_entry.get()
    workbook = check_excel_file(file_path)
    cap_nhat(workbook, ten_hang_hoa, ma_san_pham, gia_tien_nhap, so_luong, gia_ban)
    workbook.save(file_path)
    # Adjust excel to fit the text input
    adjust_excel_format(workbook)
    workbook.save(file_path)
    # Tính số tiền tổng nhập vào hàng tháng:
    workbook.close()

def xuat_hang(file_path,
        xk_ten_hang_hoa_entry,
        xk_ma_hang_hoa_entry,
        xk_so_luong_ban_ra_entry):
    xk_ten_hang_hoa = xk_ten_hang_hoa_entry.get()
    xk_ma_san_pham = xk_ma_hang_hoa_entry.get()
    xk_so_luong = xk_so_luong_ban_ra_entry.get()

    workbook = check_excel_file(file_path)
    cap_nhat_ban(workbook, xk_ten_hang_hoa, xk_ma_san_pham, xk_so_luong)
    workbook.save(file_path)
    # Adjust excel to fit the text input
    adjust_excel_format(workbook)
    workbook.save(file_path)
    # Tính số tiền tổng nhập vào hàng tháng:
    workbook.close()