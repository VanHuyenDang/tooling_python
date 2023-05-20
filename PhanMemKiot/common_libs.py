import tkinter as tk
from tkinter import ttk
import math
import openpyxl
import os.path
import numpy as np

def xuly_tinh_toan_excel(file_path):
    workbook = check_excel_file(file_path)
    sheet = workbook.active
    tong_tien_nhap_sp = 0
    doanh_thu_hang_thang = 0
    if (sheet.max_row < 2):
        print("File excel trống, không tính được")
    else:
        start_row = 2 
        end_row = sheet.max_row
        print(end_row)
        cot_gia_nhap_vao_sp = 0
        cot_sl_sp = 0
        cot_ten_sp = 0
        cot_ma_sp = 0
        cot_tong_tien_nhap_1sp = 0
        cot_gia_ban_ra = 0
        cot_sl_sp_ban_ra = 0

        for cell in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
            for idx, header in enumerate(cell):
                if header == 'Giá tiền nhập/1SP':
                    cot_gia_nhap_vao_sp = idx + 1
                elif header == 'Số lượng':
                    cot_sl_sp = idx + 1
                elif header == 'Tên sản phẩm':
                    cot_ten_sp = idx + 1
                elif header == 'Mã sản phẩm':
                    cot_ma_sp = idx + 1
                elif header == 'Giá bán/1SP':
                    cot_gia_ban_ra = idx + 1
                elif header == 'Tổng tiền nhập/1SP':
                    cot_tong_tien_nhap_1sp = idx + 1
                elif header == 'Số SP bán ra':
                    cot_sl_sp_ban_ra = idx + 1
        # print(cot_gia_nhap_vao_sp, cot_sl_sp, cot_tong_tien_nhap_1sp)
        for row in range(start_row, end_row + 1):
            sl_sp_nhap = (sheet.cell(row=row, column=cot_sl_sp)).value
            gia_nhap_sp = (sheet.cell(row=row, column=cot_gia_nhap_vao_sp)).value
            sl_sp_ban_ra = (sheet.cell(row=row, column=cot_sl_sp_ban_ra)).value
            gia_sp_ban_ra = (sheet.cell(row=row, column=cot_gia_ban_ra)).value
            if sl_sp_nhap is None: sl_sp_nhap = int(0)
            if gia_nhap_sp is None: gia_nhap_sp = float(0)
            if sl_sp_ban_ra is None: sl_sp_ban_ra = int(0)
            if gia_sp_ban_ra is None: gia_sp_ban_ra = float(0)
            
            tong_tien_nhap_sp +=float(gia_nhap_sp)*float(sl_sp_nhap)
            doanh_thu_hang_thang +=float(gia_sp_ban_ra)*float(sl_sp_ban_ra)
            sheet.cell(row=row, column=cot_tong_tien_nhap_1sp, value=float(gia_nhap_sp)*float(sl_sp_nhap))
    workbook.save(file_path)
    return tong_tien_nhap_sp, doanh_thu_hang_thang


def adjust_excel_format(workbook):
    # sheet = check_excel_file(file_path)
    sheet = workbook.active
    # Autofit column width to fit the content
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_letter].width = adjusted_width

def check_excel_file(file_path):
    # Check if the file exists
    if os.path.isfile(file_path):
        # Load the existing workbook
        workbook = openpyxl.load_workbook(file_path)
        # Select the active sheet
        sheet = workbook.active
    else:
        # Create a new workbook
        workbook = openpyxl.Workbook()
        # Set the workbook title
        workbook.title = 'Bang_Tong_Hop.xlsx'
        # Select the active sheet
        sheet = workbook.active
        # Write the header to the sheet
        sheet.append(['Tên sản phẩm', 'Mã sản phẩm', 'Giá tiền nhập/1SP', 'Giá bán/1SP', 'Số lượng', 'Tổng tiền nhập/1SP', 'Số SP bán ra'])
     
    # Save the workbook
    workbook.save(file_path)
    return workbook

def scan_excel(file_path):
    workbook = check_excel_file(file_path)
    sheet = workbook.active
    thong_tin_ten_sp = []
    thong_tin_sl_sp = []
    thong_tin_gia_sp = []
    thong_tin_ma_sp = []
    if (sheet.max_row < 2):
        thong_tin_ten_sp.append('Không có danh mục SP')
        thong_tin_ma_sp.append('Không có mã SP')
        thong_tin_gia_sp.append(0)
        thong_tin_sl_sp.append(0)
    else:
        # Get the range of cells where the cost and number of products are located
        start_row = 2  # Assuming values start from the second row
        end_row = sheet.max_row
        # Find cost (Giá/1SP) and quantity (Số lượng) cột
        cot_gia_sp = None
        cot_sl_sp = None
        cot_ten_sp = None
        cot_ma_sp = None
        cot_tong_tien = None
        for cell in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
            for idx, header in enumerate(cell):
                if header == 'Giá tiền nhập/1SP':
                    cot_gia_sp = idx + 1
                if header == 'Số lượng':
                    cot_sl_sp = idx + 1  
                if header == 'Tên sản phẩm':
                    cot_ten_sp = idx + 1
                if header == 'Mã sản phẩm':
                    cot_ma_sp = idx + 1
                if header == 'Tổng tiền nhập/1SP':
                    cot_tong_tien = idx + 1

        for row in range(start_row, end_row + 1):
            ten_sp = (sheet.cell(row=row, column=cot_ten_sp)).value
            ma_sp = (sheet.cell(row=row, column=cot_ma_sp)).value
            sl_sp = (sheet.cell(row=row, column=cot_sl_sp)).value
            gia_sp = (sheet.cell(row=row, column=cot_gia_sp)).value
            thong_tin_ten_sp.append(ten_sp)
            thong_tin_ma_sp.append(ma_sp)
            thong_tin_gia_sp.append(gia_sp)
            thong_tin_sl_sp.append(sl_sp)
    return thong_tin_ten_sp, thong_tin_ma_sp, thong_tin_gia_sp, thong_tin_sl_sp

def scan_excel_to_show(file_path, ma_sp_tim, ten_sp_tim):
    sl_sp=0
    gia_nhap_sp=0
    gia_ban_ra=0
    sl_sp_ban_ra=0
    sl_sp_ton_kho=0
    workbook = check_excel_file(file_path)
    sheet = workbook.active
    if (sheet.max_row < 2):
        print("File excel chưa cập nhật")
    else:
        start_row = 2  
        end_row = sheet.max_row
        cot_gia_nhap_sp = None
        cot_sl_sp = None
        cot_ten_sp = None
        cot_ma_sp = None
        cot_tong_tien = None
        cot_gia_ban_ra = None
        cot_sl_sp_ban_ra = None
        for cell in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
            for idx, header in enumerate(cell):
                if header == 'Giá tiền nhập/1SP':
                    cot_gia_sp = idx + 1
                elif header == 'Số lượng':
                    cot_sl_sp = idx + 1  
                elif header == 'Tên sản phẩm':
                    cot_ten_sp = idx + 1
                elif header == 'Mã sản phẩm':
                    cot_ma_sp = idx + 1
                elif header == 'Giá bán/1SP':
                    cot_gia_ban_ra = idx + 1
                elif header == 'Tổng tiền nhập/1SP':
                    cot_tong_tien = idx + 1
                elif header == 'Số SP bán ra':
                    cot_sl_sp_ban_ra = idx + 1

        for row in range(start_row, end_row + 1):
            if((sheet.cell(row=row, column=cot_ten_sp)).value == ten_sp_tim or
             (sheet.cell(row=row, column=cot_ma_sp)).value == ma_sp_tim):
                sl_sp = (sheet.cell(row=row, column=cot_sl_sp)).value
                gia_nhap_sp = (sheet.cell(row=row, column=cot_gia_sp)).value
                gia_ban_ra = (sheet.cell(row=row, column=cot_gia_ban_ra)).value
                sl_sp_ban_ra = (sheet.cell(row=row, column=cot_sl_sp_ban_ra)).value
                if sl_sp is None: sl_sp = int(0)
                if sl_sp_ban_ra is None: sl_sp_ban_ra = int(0)
                sl_sp_ton_kho = int(sl_sp) - int(sl_sp_ban_ra)
                break
            else:
                sl_sp = 0
                gia_nhap_sp = 0
                gia_ban_ra = 0
                sl_sp_ban_ra = 0
                sl_sp_ton_kho = 0
    return sl_sp, gia_nhap_sp, gia_ban_ra, sl_sp_ban_ra, sl_sp_ton_kho